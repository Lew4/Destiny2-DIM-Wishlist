#!/usr/bin/env python3
"""
DIM 中文推荐表转愿望单工具。

输入中文 Excel/CSV 推荐表 + Bungie manifest SQLite/zip content，输出 DIM wishlist txt。

核心规则：
- 武器列默认识别：名字 / 武器 / name / weapon
- perk 列默认识别：列名包含 Perk / perk / PERK
- 单元格多行 perk 自动展开组合
- 同名多版本武器会全部生成，并分别读取各自 socket
- perk 按“武器版本 + 动态 socket 映射 + 名称”解析，不做全局 hash 套用
- 未匹配武器/perk 输出到 unresolved CSV
"""
from __future__ import annotations

import csv
import itertools
import json
import os
import re
import sqlite3
import sys
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

UINT32 = 2 ** 32

# =========================
# 固定配置区：只改这里
# =========================
# 输入推荐表，支持 .csv / .xlsx。
INPUT_PATH = r"./命运2-凯旋丰碑全种类武器推荐-Sheet1.csv"

# Bungie manifest 文件。可以是刚下载的外层 .content 压缩包，也可以是解压后的 SQLite。
MANIFEST_PATH = r"./world_sql_content_22b6eb96bbcaa631746b584b52bcc2a6.content"

# Excel 工作表名；CSV 无效。None 表示读取第一个 sheet。
SHEET_NAME = None

# 输出目录。建议单独放一个文件夹，避免覆盖旧结果。
OUTPUT_DIR = r"./dim_wishlist_output"

# 输出文件名。
OUT_WISHLIST = "dim_wishlist_resolved.txt"
OUT_UNRESOLVED = "dim_wishlist_unresolved.csv"
OUT_EXTRACTED = "dim_wishlist_extracted.csv"
OUT_RESOLVED_AUDIT = "dim_wishlist_resolved_audit.csv"

# manifest 解压缓存目录。
CACHE_DIR = "./.manifest_cache"

# 武器同名候选处理策略：
# - "single": 同名查到多个武器 hash 时，只使用一个。
# - "all": 同名查到多个武器 hash 时，全部写入，并按每个 item hash 独立分组输出。推荐默认。
WEAPON_VERSION_MODE = "all"

# 同名旧版本缺少部分推荐 perk 时的处理：
# - "drop_unsupported": 该版本只保留实际支持的 perk，仍生成降级愿望单。推荐默认。
# - "strict": 只要有一个推荐 perk 不支持，就跳过该版本。
VERSION_PERK_POLICY = "drop_unsupported"

# 武器 hash 手动覆盖。用于 manifest 查出多个同名候选时指定准确版本。
# 写法示例：WEAPON_HASH_OVERRIDES = {"Yeartide Apex": [3293207827], "毫不迟疑": [123456789]}
WEAPON_HASH_OVERRIDES: Dict[str, List[int]] = {}

# 输出所有武器名匹配到的候选 hash，方便检查为什么组合数翻倍。
OUT_WEAPON_CANDIDATES = "dim_wishlist_weapon_candidates.csv"
OUT_PERK_CANDIDATES = "dim_wishlist_perk_candidates.csv"


WEAPON_COL_ALIASES = {"名字", "名称", "武器", "武器名", "name", "weapon", "item"}
NOTE_COL_ALIASES = {"注释", "备注", "说明", "notes", "note", "comment"}
TIER_COL_ALIASES = {"tier", "等级", "评级"}
RANK_COL_ALIASES = {"rank", "排序", "排名"}
EXCLUDE_PERK_HEADERS = {"原始特性", "起源特性", "origin trait", "origin"}

# 常见中文全角/半角分隔符。换行最重要。
SPLIT_RE = re.compile(r"[\n\r/／、，,;；|｜]+")

# 输入表四个固定列在武器随机词条 socket 中的位置。
# 0/1/2/3 分别对应 DIM 中写出的第1/2/3/4个 perk hash。
SLOT_TO_ROLL_INDEX = {
    "slot_1_barrel": 0,
    "slot_2_magazine": 1,
    "slot_3_trait": 2,
    "slot_4_trait": 3,
}


def norm_name(s: Any) -> str:
    if s is None:
        return ""
    text = str(s).strip()
    text = text.replace("\u3000", " ")
    text = text.lower()
    text = re.sub(r"[\s\-_'\"“”‘’·•:：()（）\[\]【】{}<>《》]+", "", text)
    return text


def clean_text(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()




def make_unique_headers(raw_headers: Sequence[Any]) -> List[str]:
    """保留重复列名。CSV/Excel 里常见多个 Perk 列，直接 DictReader 会覆盖前面的 Perk。"""
    counts: Dict[str, int] = {}
    headers: List[str] = []
    for i, h in enumerate(raw_headers):
        base = clean_text(h) or f"col_{i+1}"
        counts[base] = counts.get(base, 0) + 1
        if counts[base] == 1:
            headers.append(base)
        else:
            headers.append(f"{base}__{counts[base]}")
    return headers


def looks_like_header(row: Sequence[Any]) -> bool:
    vals = [clean_text(x) for x in row]
    has_weapon = any(norm_name(v) in {norm_name(a) for a in WEAPON_COL_ALIASES} for v in vals)
    has_perk = any("perk" in v.lower() for v in vals)
    return has_weapon and has_perk


def row_to_dict(headers: Sequence[str], raw: Sequence[Any]) -> Dict[str, Any]:
    return {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))}


def split_options(cell: Any) -> List[str]:
    text = clean_text(cell)
    if not text:
        return []
    # 去掉项目符号和多余空格。
    parts = [p.strip(" \t-—*·•") for p in SPLIT_RE.split(text)]
    return [p for p in parts if p]


def to_dim_hash(value: Any) -> int:
    """SQLite id / raw manifest hash -> DIM 使用的 uint32 hash。

    Bungie manifest 的 SQLite 表 id 常用 signed int32 存储；
    DIM wishlist 必须使用 unsigned uint32。
    例如 SQLite id = -1001759469，对应 DIM hash = 3293207827。
    """
    x = int(value)
    if x < 0:
        x += UINT32
    return x


def to_sql_id(value: Any) -> int:
    """DIM 使用的 uint32 hash -> SQLite 查询用 signed int32 id。

    手动写 SQL 时，如果 hash > 2147483647，需要减去 4294967296。
    例如 DIM hash = 3293207827，对应 SQLite id = -1001759469。
    """
    x = int(value)
    if x >= 2 ** 31:
        x -= UINT32
    return x


def sql_quote_identifier(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def ensure_sqlite_manifest(path: Path, workdir: Optional[Path] = None) -> Path:
    """返回真正可查询的 SQLite 数据库路径。

    Bungie 的 .content 常见是 zip；有时 sqlite3 shell 直接打开只看到 zip 表。
    这里优先按 zip 解压；如果不是 zip，就直接返回原路径。
    """
    path = path.expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"manifest 不存在: {path}")

    if zipfile.is_zipfile(path):
        out_dir = workdir or Path(tempfile.mkdtemp(prefix="bungie_manifest_"))
        out_dir.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(path, "r") as zf:
            names = [n for n in zf.namelist() if not n.endswith("/")]
            if not names:
                raise RuntimeError(f"zip manifest 为空: {path}")
            zf.extractall(out_dir)
            # 通常只有一个文件。
            extracted = out_dir / names[0]
            return extracted.resolve()
    return path


@dataclass
class InvItem:
    hash: int
    sql_id: int
    name: str
    item_type: Optional[int]
    item_type_display: str
    item_type_and_tier_display: str
    tier_type_name: str
    plug_category_identifier: str
    has_plug: bool
    json_obj: Dict[str, Any]


def load_inventory_items(db_path: Path) -> List[InvItem]:
    con = sqlite3.connect(str(db_path))
    cur = con.cursor()
    tables = [r[0] for r in cur.execute("SELECT name FROM sqlite_master WHERE type='table'")]
    if "DestinyInventoryItemDefinition" not in tables:
        raise RuntimeError(
            "没有找到 DestinyInventoryItemDefinition。"
            "如果 .tables 只看到 zip，说明你传入的是外层压缩文件；脚本通常会自动解压，"
            "但也可以手动 unzip 后再传入解压出的文件。"
        )

    rows = cur.execute("SELECT id, json FROM DestinyInventoryItemDefinition").fetchall()
    items: List[InvItem] = []
    for raw_id, raw_json in rows:
        try:
            obj = json.loads(raw_json)
        except Exception:
            continue
        name = obj.get("displayProperties", {}).get("name", "") or ""
        if not name:
            continue
        plug_obj = obj.get("plug") or {}
        # SQLite 的 id 可能是 signed int32；JSON 内部 hash 一般是 DIM 使用的 unsigned hash。
        # 优先使用 JSON hash，避免手动查 SQL 时的 signed/unsigned 混淆。
        dim_hash = obj.get("hash")
        if dim_hash is None:
            dim_hash = to_dim_hash(raw_id)
        items.append(
            InvItem(
                hash=to_dim_hash(dim_hash),
                sql_id=int(raw_id),
                name=name,
                item_type=obj.get("itemType"),
                item_type_display=obj.get("itemTypeDisplayName", "") or "",
                item_type_and_tier_display=obj.get("itemTypeAndTierDisplayName", "") or "",
                tier_type_name=(obj.get("inventory") or {}).get("tierTypeName", "") or "",
                plug_category_identifier=plug_obj.get("plugCategoryIdentifier", "") or "",
                has_plug=bool(obj.get("plug")),
                json_obj=obj,
            )
        )
    con.close()
    return items



def load_plug_sets(db_path: Path) -> Dict[int, List[int]]:
    """读取 DestinyPlugSetDefinition，返回 plugSetHash -> plugItemHash 列表。

    SQLite id 可能是 signed int32；字典键和 plug hash 统一转成 DIM 使用的 uint32。
    """
    con = sqlite3.connect(str(db_path))
    cur = con.cursor()
    tables = {r[0] for r in cur.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    if "DestinyPlugSetDefinition" not in tables:
        con.close()
        raise RuntimeError("manifest 中没有 DestinyPlugSetDefinition，无法按武器 socket 解析 perk。")

    out: Dict[int, List[int]] = {}
    for raw_id, raw_json in cur.execute("SELECT id, json FROM DestinyPlugSetDefinition"):
        try:
            obj = json.loads(raw_json)
        except Exception:
            continue
        set_hash = to_dim_hash(obj.get("hash", raw_id))
        hashes: List[int] = []
        for entry in obj.get("reusablePlugItems", []) or []:
            h = entry.get("plugItemHash")
            if h not in (None, 0):
                hashes.append(to_dim_hash(h))
        # 保序去重。
        out[set_hash] = list(dict.fromkeys(hashes))
    con.close()
    return out

class ManifestIndex:
    def __init__(self, items: List[InvItem], plug_sets: Dict[int, List[int]]):
        self.items = items
        self.plug_sets = plug_sets
        self.by_name: Dict[str, List[InvItem]] = {}
        self.by_hash: Dict[int, InvItem] = {}
        self._roll_socket_cache: Dict[int, List[Dict[str, Any]]] = {}
        self._slot_socket_mapping_cache: Dict[Tuple[int, Tuple[Tuple[str, Tuple[str, ...]], ...]], Dict[str, Dict[str, Any]]] = {}
        self._perk_resolution_cache: Dict[Tuple[int, str, str, Tuple[Tuple[str, Tuple[str, ...]], ...]], Tuple[Optional[InvItem], Dict[str, Any]]] = {}
        for it in items:
            self.by_name.setdefault(norm_name(it.name), []).append(it)
            self.by_hash[it.hash] = it

    def find_weapons(self, name: str) -> List[InvItem]:
        key = norm_name(name)
        exact = self.by_name.get(key, [])
        weapons = [x for x in exact if x.item_type == 3]
        if weapons:
            return sorted(weapons, key=lambda x: x.hash)
        fuzzy = [
            x for k, vals in self.by_name.items()
            if key and (key in k or k in key)
            for x in vals if x.item_type == 3
        ]
        return sorted(fuzzy, key=lambda x: (x.name, x.hash))

    @staticmethod
    def _is_enhanced_perk(c: InvItem) -> bool:
        text = " ".join([
            c.name,
            c.item_type_display,
            c.item_type_and_tier_display,
            c.tier_type_name,
        ]).lower()
        return ("强化" in text) or ("enhanced" in text)

    @staticmethod
    def _is_normal_trait(c: InvItem) -> bool:
        tt = c.item_type_and_tier_display.lower()
        is_trait_name = c.item_type_display in {"特性", "Trait"} or " trait" in tt or tt.endswith("trait")
        is_normal_tier = c.tier_type_name in {"普通", "Common"} or "普通" in tt or "common" in tt
        return is_trait_name and is_normal_tier and not ManifestIndex._is_enhanced_perk(c)

    @staticmethod
    def _perk_rank(c: InvItem) -> Tuple[int, int, int, int, int]:
        if ManifestIndex._is_normal_trait(c):
            return (0, 0, 0, 0, c.hash)
        if c.item_type == 19 and c.has_plug and not ManifestIndex._is_enhanced_perk(c):
            return (1, 0, 0, 0, c.hash)
        if c.item_type == 19 and c.has_plug:
            return (2, 0, 0, 0, c.hash)
        if c.item_type == 19:
            return (3, 0, 0, 0, c.hash)
        return (9, 9, 9, 9, c.hash)

    @staticmethod
    def _rank_perk_candidates(cands: List[InvItem]) -> List[InvItem]:
        seen = set()
        out: List[InvItem] = []
        for c in sorted(cands, key=ManifestIndex._perk_rank):
            if c.hash not in seen:
                out.append(c)
                seen.add(c.hash)
        return out

    def find_perk_hashes(self, name: str) -> List[InvItem]:
        """仅用于候选报告；正式生成不再使用全局名称匹配。"""
        exact = self.by_name.get(norm_name(name), [])
        valid = [x for x in exact if x.item_type == 19 and x.has_plug]
        return self._rank_perk_candidates(valid)

    def _collect_socket_hashes(self, socket_entry: Dict[str, Any]) -> List[int]:
        hashes: List[int] = []
        for key in ("randomizedPlugSetHash", "reusablePlugSetHash"):
            set_hash = socket_entry.get(key)
            if set_hash not in (None, 0):
                hashes.extend(self.plug_sets.get(to_dim_hash(set_hash), []))
        for entry in socket_entry.get("reusablePlugItems", []) or []:
            h = entry.get("plugItemHash")
            if h not in (None, 0):
                hashes.append(to_dim_hash(h))
        return list(dict.fromkeys(hashes))

    def get_weapon_roll_sockets(self, weapon_hash: int) -> List[Dict[str, Any]]:
        """返回武器的四个随机 roll socket，顺序即 DIM perk 顺序。\n\n        首选 randomizedPlugSetHash。它专门指向该 socket 的随机词条集合。\n        极少数定义不足四个时，才用其他可复用 plug socket 补齐，并排除框架、起源特性、模组等。\n        """
        weapon_hash = to_dim_hash(weapon_hash)
        if weapon_hash in self._roll_socket_cache:
            return self._roll_socket_cache[weapon_hash]

        weapon = self.by_hash.get(weapon_hash)
        if not weapon:
            return []
        entries = (weapon.json_obj.get("sockets") or {}).get("socketEntries", []) or []
        randomized: List[Dict[str, Any]] = []
        fallback: List[Dict[str, Any]] = []

        for socket_index, entry in enumerate(entries):
            candidate_hashes = self._collect_socket_hashes(entry)
            candidates = [self.by_hash[h] for h in candidate_hashes if h in self.by_hash]
            candidates = [c for c in candidates if c.item_type == 19 and c.has_plug]
            if not candidates:
                continue

            categories = {c.plug_category_identifier.lower() for c in candidates if c.plug_category_identifier}
            info = {
                "socket_index": socket_index,
                "socket_type_hash": to_dim_hash(entry.get("socketTypeHash", 0)),
                "randomized_plug_set_hash": to_dim_hash(entry.get("randomizedPlugSetHash", 0)) if entry.get("randomizedPlugSetHash") else 0,
                "reusable_plug_set_hash": to_dim_hash(entry.get("reusablePlugSetHash", 0)) if entry.get("reusablePlugSetHash") else 0,
                "candidate_hashes": [c.hash for c in candidates],
                "candidates": candidates,
                "categories": sorted(categories),
            }
            if entry.get("randomizedPlugSetHash"):
                randomized.append(info)
                continue

            joined = " ".join(categories)
            excluded_tokens = (
                "frame", "intrinsic", "origin", "masterwork", "tracker", "mod", "memento",
                "shader", "cosmetic", "deepsight", "enhancement", "crafting_plug",
            )
            if not any(token in joined for token in excluded_tokens):
                fallback.append(info)

        # socketEntries 本身有固定顺序。通常前四个 randomized socket 就是1~4号位。
        sockets = randomized[:4]
        if len(sockets) < 4:
            used = {x["socket_index"] for x in sockets}
            for info in fallback:
                if info["socket_index"] not in used:
                    sockets.append(info)
                    used.add(info["socket_index"])
                if len(sockets) >= 4:
                    break

        self._roll_socket_cache[weapon_hash] = sockets
        return sockets

    @staticmethod
    def _slot_category_bonus(slot: str, categories: Sequence[str]) -> int:
        """根据 plugCategoryIdentifier 给槽位一个很小的辅助分数。

        名称命中始终是主依据；分类只用于多个映射得分相同时的稳定选择。
        """
        text = " ".join(categories).lower()
        if slot == "slot_1_barrel":
            tokens = ("barrel", "sight", "scope", "bowstring", "string", "blade", "haft")
        elif slot == "slot_2_magazine":
            tokens = ("magazine", "battery", "arrow", "fletching", "guard", "projectile", "ammo")
        else:
            tokens = ("trait", "perk")
        return 8 if any(token in text for token in tokens) else 0

    @staticmethod
    def _slot_options_signature(slot_options: Optional[Dict[str, Sequence[str]]]) -> Tuple[Tuple[str, Tuple[str, ...]], ...]:
        slot_options = slot_options or {}
        ordered: List[Tuple[str, Tuple[str, ...]]] = []
        for slot in SLOT_TO_ROLL_INDEX:
            vals = tuple(sorted({norm_name(x) for x in slot_options.get(slot, []) if norm_name(x)}))
            if vals:
                ordered.append((slot, vals))
        return tuple(ordered)

    def infer_slot_socket_mapping(
        self,
        weapon_hash: int,
        slot_options: Optional[Dict[str, Sequence[str]]],
    ) -> Dict[str, Dict[str, Any]]:
        """根据本行全部推荐 perk，推断四个输入列对应的真实武器 socket。

        不再假设 manifest 的前四个随机 socket 永远严格对应输入四列。算法会：
        1. 统计每个输入列的推荐名称在各 socket 候选中的命中数量；
        2. 对最多四个 socket 做一对一排列；
        3. 优先选择名称命中最多的映射；
        4. 命中相同后，才使用类别与原始顺序作为稳定的次级依据。
        """
        weapon_hash = to_dim_hash(weapon_hash)
        signature = self._slot_options_signature(slot_options)
        cache_key = (weapon_hash, signature)
        if cache_key in self._slot_socket_mapping_cache:
            return self._slot_socket_mapping_cache[cache_key]

        sockets = self.get_weapon_roll_sockets(weapon_hash)
        slots = [slot for slot, vals in signature if vals]
        if not slots:
            slots = list(SLOT_TO_ROLL_INDEX)
        slots = slots[:len(sockets)]

        result: Dict[str, Dict[str, Any]] = {}
        if not sockets or not slots:
            self._slot_socket_mapping_cache[cache_key] = result
            return result

        option_map = {slot: set(vals) for slot, vals in signature}
        socket_name_sets = [
            {norm_name(c.name) for c in socket.get("candidates", []) if norm_name(c.name)}
            for socket in sockets
        ]

        best_perm: Optional[Tuple[int, ...]] = None
        best_objective: Optional[Tuple[int, int, int, int]] = None
        best_slot_scores: Dict[str, Dict[str, int]] = {}

        for perm in itertools.permutations(range(len(sockets)), len(slots)):
            total_hits = 0
            slots_with_hits = 0
            category_bonus = 0
            order_penalty = 0
            local_scores: Dict[str, Dict[str, int]] = {}
            for slot, roll_index in zip(slots, perm):
                opts = option_map.get(slot, set())
                hits = len(opts & socket_name_sets[roll_index])
                if hits > 0:
                    slots_with_hits += 1
                total_hits += hits
                cat_bonus = self._slot_category_bonus(slot, sockets[roll_index].get("categories", []))
                category_bonus += cat_bonus
                order_penalty += abs(SLOT_TO_ROLL_INDEX.get(slot, roll_index) - roll_index)
                local_scores[slot] = {
                    "name_hits": hits,
                    "category_bonus": cat_bonus,
                    "order_distance": abs(SLOT_TO_ROLL_INDEX.get(slot, roll_index) - roll_index),
                }

            # 名称命中优先级远高于类别和原顺序。
            objective = (total_hits, slots_with_hits, category_bonus, -order_penalty)
            if best_objective is None or objective > best_objective:
                best_objective = objective
                best_perm = perm
                best_slot_scores = local_scores

        if best_perm is None:
            self._slot_socket_mapping_cache[cache_key] = result
            return result

        for slot, roll_index in zip(slots, best_perm):
            socket = sockets[roll_index]
            score = best_slot_scores.get(slot, {})
            result[slot] = {
                "roll_index": roll_index,
                "socket_index": socket.get("socket_index", ""),
                "name_hits": score.get("name_hits", 0),
                "category_bonus": score.get("category_bonus", 0),
                "order_distance": score.get("order_distance", 0),
                "mapping_method": "name_coverage_assignment" if score.get("name_hits", 0) > 0 else "ordered_fallback",
            }

        self._slot_socket_mapping_cache[cache_key] = result
        return result

    def resolve_perk_for_weapon_slot(
        self,
        weapon_hash: int,
        slot: str,
        perk_name: str,
        slot_options: Optional[Dict[str, Sequence[str]]] = None,
    ) -> Tuple[Optional[InvItem], Dict[str, Any]]:
        """在指定武器版本中，先推断输入列对应 socket，再解析 perk。"""
        signature = self._slot_options_signature(slot_options)
        cache_key = (to_dim_hash(weapon_hash), slot, norm_name(perk_name), signature)
        if cache_key in self._perk_resolution_cache:
            return self._perk_resolution_cache[cache_key]

        sockets = self.get_weapon_roll_sockets(weapon_hash)
        mapping = self.infer_slot_socket_mapping(weapon_hash, slot_options)
        mapped = mapping.get(slot)
        detail: Dict[str, Any] = {
            "weapon_hash": to_dim_hash(weapon_hash),
            "slot": slot,
            "roll_index": "",
            "socket_index": "",
            "candidate_hashes": [],
            "candidate_names": [],
            "same_name_hashes": [],
            "mapping_method": "",
            "mapping_name_hits": 0,
            "reason": "",
        }
        if mapped is None:
            default_roll_index = SLOT_TO_ROLL_INDEX.get(slot)
            if default_roll_index is None:
                detail["reason"] = "unknown_input_slot"
                result = (None, detail)
                self._perk_resolution_cache[cache_key] = result
                return result
            if default_roll_index >= len(sockets):
                detail["reason"] = f"weapon_has_only_{len(sockets)}_roll_sockets"
                result = (None, detail)
                self._perk_resolution_cache[cache_key] = result
                return result
            mapped = {
                "roll_index": default_roll_index,
                "socket_index": sockets[default_roll_index].get("socket_index", ""),
                "mapping_method": "ordered_fallback",
                "name_hits": 0,
            }

        roll_index = int(mapped["roll_index"])
        if roll_index >= len(sockets):
            detail["reason"] = f"mapped_roll_index_out_of_range_{roll_index}"
            result = (None, detail)
            self._perk_resolution_cache[cache_key] = result
            return result

        socket = sockets[roll_index]
        candidates: List[InvItem] = socket["candidates"]
        detail["roll_index"] = roll_index
        detail["socket_index"] = socket["socket_index"]
        detail["candidate_hashes"] = [c.hash for c in candidates]
        detail["candidate_names"] = [c.name for c in candidates]
        detail["plug_categories"] = socket.get("categories", [])
        detail["mapping_method"] = mapped.get("mapping_method", "")
        detail["mapping_name_hits"] = mapped.get("name_hits", 0)

        key = norm_name(perk_name)
        exact = [c for c in candidates if norm_name(c.name) == key]
        detail["same_name_hashes"] = [c.hash for c in exact]
        if not exact:
            detail["reason"] = "perk_not_in_mapped_weapon_socket"
            result = (None, detail)
            self._perk_resolution_cache[cache_key] = result
            return result

        ranked = self._rank_perk_candidates(exact)
        selected = ranked[0] if ranked else None
        if not selected:
            detail["reason"] = "same_name_candidates_filtered_out"
        else:
            detail["reason"] = "resolved_from_mapped_weapon_socket"
            detail["selected_hash"] = selected.hash
        result = (selected, detail)
        self._perk_resolution_cache[cache_key] = result
        return result

def read_table(input_path: Path, sheet: Optional[str] = None) -> Tuple[List[str], List[Dict[str, Any]]]:
    suffix = input_path.suffix.lower()

    def build_from_rows(all_rows: List[List[Any]]) -> Tuple[List[str], List[Dict[str, Any]]]:
        # 你的表是“武器大类行 + 表头行 + 数据行 + 下一个武器大类行 + 表头行...”。
        # 所以不能只用第一行当表头，要在前若干行里找真正含“名字”和“Perk”的表头。
        header_idx = None
        for i, row in enumerate(all_rows[:80]):
            if looks_like_header(row):
                header_idx = i
                break
        if header_idx is None:
            preview = [[clean_text(x) for x in r] for r in all_rows[:8]]
            raise RuntimeError(
                "没有识别到表头。请确保某一行同时包含 `名字` 列和至少一个 `Perk` 列。"
                f"\n前几行内容: {preview}"
            )

        headers = make_unique_headers(all_rows[header_idx])
        rows: List[Dict[str, Any]] = []
        for raw in all_rows[header_idx + 1:]:
            vals = [clean_text(x) for x in raw]
            if not any(vals):
                continue
            # 跳过后续武器类别之间重复出现的表头行。
            if looks_like_header(raw):
                continue
            row = row_to_dict(headers, raw)
            rows.append(row)
        return headers, rows

    if suffix in {".xlsx", ".xlsm"}:
        if load_workbook is None:
            raise RuntimeError("读取 xlsx 需要 openpyxl：python3 -m pip install openpyxl")
        wb = load_workbook(str(input_path), data_only=True)
        ws = wb[sheet] if sheet else wb.active
        all_rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        return build_from_rows(all_rows)

    if suffix == ".csv":
        # 不用 csv.DictReader，因为你的文件第一行是“自动步枪”这种分类标题，
        # 而且存在多个同名 Perk 列，DictReader 会把重复列覆盖掉。
        with input_path.open("r", encoding="utf-8-sig", newline="") as f:
            all_rows = list(csv.reader(f))
        return build_from_rows(all_rows)

    raise RuntimeError(f"暂不支持输入格式: {suffix}")


def detect_columns(headers: Sequence[str]) -> Tuple[str, List[str], Optional[str], Optional[str], Optional[str]]:
    weapon_col = None
    for h in headers:
        if norm_name(h) in {norm_name(a) for a in WEAPON_COL_ALIASES}:
            weapon_col = h
            break
    if not weapon_col:
        raise RuntimeError(f"没有找到武器名列。可用列名: {headers}")

    perk_cols = []
    for h in headers:
        base_h = h.split("__", 1)[0]
        hn = norm_name(base_h)
        if base_h in EXCLUDE_PERK_HEADERS or hn in {norm_name(x) for x in EXCLUDE_PERK_HEADERS}:
            continue
        if "perk" in base_h.lower():
            perk_cols.append(h)
    if not perk_cols:
        raise RuntimeError("没有找到 Perk 列。列名需要包含 `Perk`。")

    note_col = next((h for h in headers if norm_name(h) in {norm_name(a) for a in NOTE_COL_ALIASES}), None)
    tier_col = next((h for h in headers if norm_name(h) in {norm_name(a) for a in TIER_COL_ALIASES}), None)
    rank_col = next((h for h in headers if norm_name(h) in {norm_name(a) for a in RANK_COL_ALIASES}), None)
    return weapon_col, perk_cols, note_col, tier_col, rank_col


def make_note(row: Dict[str, Any], note_col: Optional[str], tier_col: Optional[str], rank_col: Optional[str]) -> str:
    parts = []
    if tier_col and clean_text(row.get(tier_col)):
        parts.append(f"Tier {clean_text(row.get(tier_col))}")
    if rank_col and clean_text(row.get(rank_col)):
        parts.append(f"Rank {clean_text(row.get(rank_col))}")
    if note_col and clean_text(row.get(note_col)):
        parts.append(clean_text(row.get(note_col)).replace("\n", " "))
    return " | ".join(parts)


def canonical_perk_slot(col: str, index: int) -> str:
    """把原始列名固定映射为 DIM 检查用的 slot 名称。

    你的表头是 Perk, Perk, Perk 1, Perk 2。第二个重复 Perk
    会被内部重命名为 Perk__2，但语义仍然是第二列。
    """
    base = col.split("__", 1)[0].strip()
    if index == 0:
        return "slot_1_barrel"
    if index == 1:
        return "slot_2_magazine"
    if base.lower() in {"perk 1", "trait 1", "特性1", "特性 1"}:
        return "slot_3_trait"
    if base.lower() in {"perk 2", "trait 2", "特性2", "特性 2"}:
        return "slot_4_trait"
    return f"slot_{index + 1}"


def expand_rows(rows: List[Dict[str, Any]], weapon_col: str, perk_cols: List[str], note_col: Optional[str], tier_col: Optional[str], rank_col: Optional[str]) -> List[Dict[str, Any]]:
    expanded = []
    for row_idx, row in enumerate(rows, start=1):
        weapon = clean_text(row.get(weapon_col))
        if not weapon:
            continue
        # 类目行通常没有 perk。
        perk_options_by_col: List[List[str]] = []
        slot_names_by_col: List[str] = []
        raw_perk_columns: Dict[str, str] = {}
        parsed_perk_columns: Dict[str, List[str]] = {}
        for col_index, col in enumerate(perk_cols):
            raw_value = clean_text(row.get(col))
            opts = split_options(row.get(col))
            slot_name = canonical_perk_slot(col, col_index)
            if raw_value:
                raw_perk_columns[slot_name] = raw_value
            if opts:
                parsed_perk_columns[slot_name] = opts
                perk_options_by_col.append(opts)
                slot_names_by_col.append(slot_name)
        if not perk_options_by_col:
            continue
        for combo in itertools.product(*perk_options_by_col):
            perk_slots = {slot: perk for slot, perk in zip(slot_names_by_col, combo)}
            expanded.append({
                "source_row": row_idx,
                "weapon": weapon,
                "perks": [perk_slots[slot] for slot in slot_names_by_col],
                "perk_slots": perk_slots,
                "slot_order": list(slot_names_by_col),
                "notes": make_note(row, note_col, tier_col, rank_col),
                "raw_perk_columns": raw_perk_columns,
                "parsed_perk_columns": parsed_perk_columns,
            })
    return expanded


def write_extracted_report(path: Path, expanded: List[Dict[str, Any]]) -> None:
    """输出从原始推荐表识别并展开后的内容，先于 hash 匹配用于人工检查。"""
    rows = []
    for rec in expanded:
        perk_slots = rec.get("perk_slots", {}) or {}
        rows.append({
            "source_row": rec.get("source_row", ""),
            "weapon": rec.get("weapon", ""),
            "slot_1_barrel": perk_slots.get("slot_1_barrel", ""),
            "slot_2_magazine": perk_slots.get("slot_2_magazine", ""),
            "slot_3_trait": perk_slots.get("slot_3_trait", ""),
            "slot_4_trait": perk_slots.get("slot_4_trait", ""),
            "expanded_perks_in_dim_order": " / ".join(rec.get("perks", [])),
            "notes": rec.get("notes", ""),
            "raw_perk_columns_json": json.dumps(rec.get("raw_perk_columns", {}), ensure_ascii=False),
            "parsed_perk_columns_json": json.dumps(rec.get("parsed_perk_columns", {}), ensure_ascii=False),
        })
    csv_write(
        path,
        rows,
        [
            "source_row",
            "weapon",
            "slot_1_barrel",
            "slot_2_magazine",
            "slot_3_trait",
            "slot_4_trait",
            "expanded_perks_in_dim_order",
            "notes",
            "raw_perk_columns_json",
            "parsed_perk_columns_json",
        ],
    )


def csv_write(path: Path, rows: List[Dict[str, Any]], fields: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(fields))
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in fields})


def sanitize_comment(text: Any) -> str:
    """清理写入 DIM 注释行的文本，避免换行破坏 block notes。"""
    return clean_text(text).replace("\r", " ").replace("\n", " ").strip()



def describe_inv_item(c: InvItem) -> str:
    return f"{c.hash}:{c.name}:{c.item_type_display}:{c.tier_type_name}"


def build_group_header(weapon: str, notes: str) -> List[str]:
    """生成接近 Little Light / DIM 官方工具的分组注释样式。"""
    header = [f"// {sanitize_comment(weapon)} - recommended"]
    note_text = sanitize_comment(notes)
    if note_text:
        header.append(f"//notes: {note_text}")
    return header



def select_weapon_candidates(weapon: str, candidates: List[InvItem]) -> List[InvItem]:
    """根据固定配置决定同名武器候选的写入方式。

    组合数量应首先由四个 perk 槽位决定。例如 2×2×3×3=36。
    如果同一武器名在 manifest 中匹配到 2 个 item hash，而又全部写入，最终会变成 36×2=72。
    """
    if not candidates:
        return []

    # 手动覆盖优先。支持中文名或英文名，只要表里 weapon 字段能 norm 后匹配。
    wkey = norm_name(weapon)
    for k, hashes in WEAPON_HASH_OVERRIDES.items():
        if norm_name(k) == wkey:
            allowed = {int(h) for h in hashes}
            picked = [c for c in candidates if c.hash in allowed]
            if picked:
                return picked
            # 覆盖写错时不要静默回退，避免误生成。
            return []

    mode = str(WEAPON_VERSION_MODE).strip().lower()
    if mode == "all":
        return candidates
    if mode == "single":
        # 稳定选择第一个候选。候选排序已在 find_weapons 中完成。
        return candidates[:1]
    raise RuntimeError(f"未知 WEAPON_VERSION_MODE: {WEAPON_VERSION_MODE!r}，只能是 'single' 或 'all'")


def write_weapon_candidate_report(path: Path, index: ManifestIndex, expanded: List[Dict[str, Any]]) -> None:
    """输出每个武器名在 manifest 中匹配到的 item hash 候选。"""
    seen_names = []
    seen_key = set()
    for rec in expanded:
        w = rec.get("weapon", "")
        k = norm_name(w)
        if w and k not in seen_key:
            seen_names.append(w)
            seen_key.add(k)

    rows = []
    for weapon in seen_names:
        candidates = index.find_weapons(weapon)
        selected = select_weapon_candidates(weapon, candidates)
        cand_hashes = [c.hash for c in candidates]
        selected_hashes = [c.hash for c in selected]
        for c in candidates:
            rows.append({
                "weapon": weapon,
                "candidate_count": len(candidates),
                "selected_count": len(selected),
                "selected": "yes" if c.hash in selected_hashes else "no",
                "hash": c.hash,
                "sqlite_id": c.sql_id,
                "manifest_name": c.name,
                "item_type_display": c.item_type_display,
                "item_type_and_tier_display": c.item_type_and_tier_display,
                "tier_type_name": c.tier_type_name,
                "item_type": c.item_type,
                "has_plug": c.has_plug,
                "all_candidate_hashes": ";".join(str(x) for x in cand_hashes),
                "selected_hashes": ";".join(str(x) for x in selected_hashes),
            })
        if not candidates:
            rows.append({
                "weapon": weapon,
                "candidate_count": 0,
                "selected_count": 0,
                "selected": "no",
                "hash": "",
                "sqlite_id": "",
                "manifest_name": "",
                "item_type_display": "",
                "item_type_and_tier_display": "",
                "tier_type_name": "",
                "item_type": "",
                "has_plug": "",
                "all_candidate_hashes": "",
                "selected_hashes": "",
            })
    csv_write(path, rows, [
        "weapon", "candidate_count", "selected_count", "selected", "hash", "sqlite_id", "manifest_name",
        "item_type_display", "item_type_and_tier_display", "tier_type_name", "item_type", "has_plug",
        "all_candidate_hashes", "selected_hashes",
    ])


def write_perk_candidate_report(path: Path, index: ManifestIndex, expanded: List[Dict[str, Any]]) -> None:
    """按武器版本和固定槽位输出 perk 解析候选。"""
    rows: List[Dict[str, Any]] = []
    seen = set()
    for rec in expanded:
        weapon = rec.get("weapon", "")
        source_row = rec.get("source_row", "")
        weapon_candidates = select_weapon_candidates(weapon, index.find_weapons(weapon))
        perk_slots = rec.get("perk_slots", {}) or {}
        for wc in weapon_candidates:
            for slot, perk in perk_slots.items():
                key = (wc.hash, slot, norm_name(perk))
                if not perk or key in seen:
                    continue
                seen.add(key)
                selected, detail = index.resolve_perk_for_weapon_slot(
                    wc.hash, slot, perk, rec.get("parsed_perk_columns", {})
                )
                rows.append({
                    "source_row": source_row,
                    "weapon": weapon,
                    "weapon_hash": wc.hash,
                    "weapon_sqlite_id": wc.sql_id,
                    "slot": slot,
                    "roll_index": detail.get("roll_index", ""),
                    "socket_index": detail.get("socket_index", ""),
                    "requested_perk": perk,
                    "selected_hash": selected.hash if selected else "",
                    "selected_sqlite_id": selected.sql_id if selected else "",
                    "selected_name": selected.name if selected else "",
                    "selected_type": selected.item_type_display if selected else "",
                    "reason": detail.get("reason", ""),
                    "mapping_method": detail.get("mapping_method", ""),
                    "mapping_name_hits": detail.get("mapping_name_hits", 0),
                    "socket_candidate_count": len(detail.get("candidate_hashes", [])),
                    "socket_candidate_hashes": ";".join(str(x) for x in detail.get("candidate_hashes", [])),
                    "socket_candidate_names": ";".join(detail.get("candidate_names", [])),
                    "same_name_hashes_in_socket": ";".join(str(x) for x in detail.get("same_name_hashes", [])),
                    "plug_categories": ";".join(detail.get("plug_categories", [])),
                })
    csv_write(path, rows, [
        "source_row", "weapon", "weapon_hash", "weapon_sqlite_id", "slot", "roll_index", "socket_index",
        "requested_perk", "selected_hash", "selected_sqlite_id", "selected_name", "selected_type", "reason",
        "mapping_method", "mapping_name_hits", "socket_candidate_count", "socket_candidate_hashes", "socket_candidate_names",
        "same_name_hashes_in_socket", "plug_categories",
    ])

def collapse_expanded_recommendations(expanded: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """把笛卡尔积展开结果还原成每个原始推荐行一条记录。

    版本兼容性必须基于某一行的全部可选 perk 计算，不能逐组合处理；
    否则一个旧版本缺少一个 perk 时，会把整组其他兼容组合也一起丢掉。
    """
    unique: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
    for rec in expanded:
        parsed = rec.get("parsed_perk_columns", {}) or {}
        parsed_sig = tuple(
            (slot, tuple(parsed.get(slot, []) or []))
            for slot in rec.get("slot_order", []) or []
        )
        key = (
            rec.get("source_row"), rec.get("weapon"), rec.get("notes"), parsed_sig,
        )
        if key not in unique:
            unique[key] = {
                "source_row": rec.get("source_row"),
                "weapon": rec.get("weapon", ""),
                "notes": rec.get("notes", ""),
                "slot_order": list(rec.get("slot_order", []) or []),
                "parsed_perk_columns": {
                    slot: list(parsed.get(slot, []) or [])
                    for slot in rec.get("slot_order", []) or []
                },
            }
    return list(unique.values())


def build_wishlist(index: ManifestIndex, expanded: List[Dict[str, Any]]) -> Tuple[List[str], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """按武器版本生成愿望单，并对旧版本执行 perk 降级。

    full 版本：该推荐行中的所有 perk 都能在该版本对应 socket 中找到，输出完整笛卡尔积。
    partial 版本：只保留该版本实际支持的 perk；不支持的 perk 写入 unresolved，但不阻止
    其他兼容 perk 组合生成。若某个槽位没有任何兼容 perk，则该槽位从 DIM 规则中省略。
    DIM 官方格式允许 `perks` 包含一个或多个 hash。
    """
    policy = str(VERSION_PERK_POLICY).strip().lower()
    if policy not in {"drop_unsupported", "strict"}:
        raise RuntimeError(
            f"未知 VERSION_PERK_POLICY: {VERSION_PERK_POLICY!r}，"
            "只能是 'drop_unsupported' 或 'strict'"
        )

    lines = [
        "title:Converted Chinese Weapon Wishlist",
        "description:Generated with weapon-specific dynamic socket mapping and version fallback.",
        "",
    ]
    unresolved: List[Dict[str, Any]] = []
    resolved_audit: List[Dict[str, Any]] = []
    groups: Dict[Tuple[int, str, str, int, str, str], List[str]] = {}
    seen_lines_by_group: Dict[Tuple[int, str, str, int, str, str], set] = {}

    source_recommendations = collapse_expanded_recommendations(expanded)

    for rec in source_recommendations:
        row_id = rec["source_row"]
        weapon = rec["weapon"]
        notes = rec.get("notes", "")
        slot_order = rec.get("slot_order", []) or []
        slot_options = rec.get("parsed_perk_columns", {}) or {}

        all_weapon_candidates = index.find_weapons(weapon)
        weapon_candidates = select_weapon_candidates(weapon, all_weapon_candidates)
        if not all_weapon_candidates:
            unresolved.append({
                "source_row": row_id, "type": "weapon", "name": weapon,
                "reason": "weapon_not_found", "weapon": weapon,
                "generated": "no", "version_status": "skipped",
                "perks": " / ".join(
                    p for slot in slot_order for p in slot_options.get(slot, [])
                ),
            })
            continue
        if not weapon_candidates:
            unresolved.append({
                "source_row": row_id, "type": "weapon", "name": weapon,
                "reason": "weapon_override_or_selection_empty", "weapon": weapon,
                "generated": "no", "version_status": "skipped",
                "perks": " / ".join(
                    p for slot in slot_order for p in slot_options.get(slot, [])
                ),
            })
            continue

        for wc in weapon_candidates:
            supported_by_slot: Dict[str, List[InvItem]] = {}
            details_by_slot_name: Dict[Tuple[str, str], Dict[str, Any]] = {}
            dropped: List[Tuple[str, str, Dict[str, Any]]] = []
            requested_count = 0

            # 同一个显示名称可能在 manifest 中有多个 hash，因此必须针对当前版本、当前 socket 逐项解析。
            for slot in slot_order:
                selected_items: List[InvItem] = []
                seen_hashes = set()
                for perk_name in slot_options.get(slot, []) or []:
                    requested_count += 1
                    selected, detail = index.resolve_perk_for_weapon_slot(
                        wc.hash, slot, perk_name, slot_options
                    )
                    details_by_slot_name[(slot, perk_name)] = detail
                    if selected is None:
                        dropped.append((slot, perk_name, detail))
                        continue
                    if selected.hash not in seen_hashes:
                        selected_items.append(selected)
                        seen_hashes.add(selected.hash)
                supported_by_slot[slot] = selected_items

            supported_count = sum(len(v) for v in supported_by_slot.values())
            version_status = "full" if not dropped else "partial"

            if policy == "strict" and dropped:
                for slot, perk_name, detail in dropped:
                    unresolved.append({
                        "source_row": row_id,
                        "type": "perk",
                        "name": perk_name,
                        "reason": detail.get("reason", "perk_not_in_mapped_weapon_socket"),
                        "weapon": weapon,
                        "weapon_hash": wc.hash,
                        "slot": slot,
                        "socket_index": detail.get("socket_index", ""),
                        "mapping_method": detail.get("mapping_method", ""),
                        "mapping_name_hits": detail.get("mapping_name_hits", 0),
                        "generated": "no",
                        "version_status": "skipped_strict",
                        "requested_option_count": requested_count,
                        "supported_option_count": supported_count,
                        "perks": " / ".join(
                            p for s in slot_order for p in slot_options.get(s, [])
                        ),
                        "same_name_candidate_hashes": ";".join(str(x) for x in detail.get("same_name_hashes", [])),
                        "same_name_candidate_types": "",
                        "socket_candidate_hashes": ";".join(str(x) for x in detail.get("candidate_hashes", [])),
                        "socket_candidate_names": ";".join(detail.get("candidate_names", [])),
                    })
                continue

            # 记录被旧版本舍弃的 perk，但 generated=yes 表示该版本仍生成了兼容子集。
            for slot, perk_name, detail in dropped:
                unresolved.append({
                    "source_row": row_id,
                    "type": "perk",
                    "name": perk_name,
                    "reason": "perk_dropped_for_incompatible_version",
                    "original_reason": detail.get("reason", "perk_not_in_mapped_weapon_socket"),
                    "weapon": weapon,
                    "weapon_hash": wc.hash,
                    "slot": slot,
                    "socket_index": detail.get("socket_index", ""),
                    "mapping_method": detail.get("mapping_method", ""),
                    "mapping_name_hits": detail.get("mapping_name_hits", 0),
                    "generated": "yes",
                    "version_status": version_status,
                    "requested_option_count": requested_count,
                    "supported_option_count": supported_count,
                    "perks": " / ".join(
                        p for s in slot_order for p in slot_options.get(s, [])
                    ),
                    "same_name_candidate_hashes": ";".join(str(x) for x in detail.get("same_name_hashes", [])),
                    "same_name_candidate_types": "",
                    "socket_candidate_hashes": ";".join(str(x) for x in detail.get("candidate_hashes", [])),
                    "socket_candidate_names": ";".join(detail.get("candidate_names", [])),
                })

            active_slots = [slot for slot in slot_order if supported_by_slot.get(slot)]
            if not active_slots:
                unresolved.append({
                    "source_row": row_id,
                    "type": "version",
                    "name": weapon,
                    "reason": "version_has_no_compatible_recommended_perks",
                    "weapon": weapon,
                    "weapon_hash": wc.hash,
                    "generated": "no",
                    "version_status": "skipped",
                    "requested_option_count": requested_count,
                    "supported_option_count": 0,
                    "perks": " / ".join(
                        p for s in slot_order for p in slot_options.get(s, [])
                    ),
                })
                continue

            dropped_names = [perk for _slot, perk, _detail in dropped]
            combo_lists = [supported_by_slot[slot] for slot in active_slots]
            group_key = (
                int(row_id), sanitize_comment(weapon), sanitize_comment(notes),
                int(wc.hash), sanitize_comment(wc.name), version_status,
            )
            groups.setdefault(group_key, [])
            seen_lines_by_group.setdefault(group_key, set())

            for combo in itertools.product(*combo_lists):
                selected_by_slot = {slot: item for slot, item in zip(active_slots, combo)}
                perk_hashes = [selected_by_slot[slot].hash for slot in active_slots]
                if not perk_hashes:
                    continue
                perk_part = ",".join(str(x) for x in perk_hashes)
                line = f"dimwishlist:item={wc.hash}&perks={perk_part}"
                if line in seen_lines_by_group[group_key]:
                    continue
                groups[group_key].append(line)
                seen_lines_by_group[group_key].add(line)

                audit: Dict[str, Any] = {
                    "source_row": row_id,
                    "weapon": weapon,
                    "weapon_hash": wc.hash,
                    "weapon_sqlite_id": wc.sql_id,
                    "manifest_name": wc.name,
                    "version_status": version_status,
                    "included_slots": ";".join(active_slots),
                    "dropped_perks": ";".join(dropped_names),
                    "requested_option_count": requested_count,
                    "supported_option_count": supported_count,
                    "dim_line": line,
                }
                for slot in SLOT_TO_ROLL_INDEX:
                    requested = slot_options.get(slot, []) or []
                    supported = supported_by_slot.get(slot, []) or []
                    selected = selected_by_slot.get(slot)
                    selected_detail = (
                        details_by_slot_name.get((slot, selected.name), {}) if selected else {}
                    )
                    audit[f"{slot}_requested_options"] = ";".join(requested)
                    audit[f"{slot}_supported_options"] = ";".join(x.name for x in supported)
                    audit[slot] = selected.name if selected else ""
                    audit[f"{slot}_hash"] = selected.hash if selected else ""
                    audit[f"{slot}_sqlite_id"] = selected.sql_id if selected else ""
                    audit[f"{slot}_type"] = selected.item_type_display if selected else ""
                    audit[f"{slot}_socket_index"] = selected_detail.get("socket_index", "")
                resolved_audit.append(audit)

    wrote_any = False
    for group_key, group_lines in groups.items():
        if not group_lines:
            continue
        _row_id, weapon_name, notes, weapon_hash, manifest_name, version_status = group_key
        if wrote_any:
            lines.append("")
        header_weapon = weapon_name
        if manifest_name and norm_name(manifest_name) != norm_name(weapon_name):
            header_weapon = f"{weapon_name} / {manifest_name}"
        if version_status == "partial":
            header_weapon = f"{header_weapon} [兼容子集]"
        lines.extend(build_group_header(f"{header_weapon} [{weapon_hash}]", notes))
        lines.extend(group_lines)
        wrote_any = True

    unique_unresolved: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
    for row in unresolved:
        key = (
            row.get("source_row"), row.get("type"), row.get("weapon"), row.get("weapon_hash"),
            row.get("slot"), row.get("name"), row.get("reason"), row.get("socket_index"),
        )
        unique_unresolved.setdefault(key, row)

    return lines, list(unique_unresolved.values()), resolved_audit

def main() -> int:
    input_path = Path(INPUT_PATH).expanduser().resolve()
    manifest_path = Path(MANIFEST_PATH).expanduser().resolve()
    output_dir = Path(OUTPUT_DIR).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    out_path = output_dir / OUT_WISHLIST
    unresolved_path = output_dir / OUT_UNRESOLVED
    extracted_path = output_dir / OUT_EXTRACTED
    resolved_audit_path = output_dir / OUT_RESOLVED_AUDIT
    weapon_candidates_path = output_dir / OUT_WEAPON_CANDIDATES
    perk_candidates_path = output_dir / OUT_PERK_CANDIDATES
    cache_dir = Path(CACHE_DIR).expanduser().resolve()

    print("[CONFIG]")
    print(f"      input: {input_path}")
    print(f"      manifest: {manifest_path}")
    print(f"      output_dir: {output_dir}")
    print(f"      weapon_version_mode: {WEAPON_VERSION_MODE}")
    print(f"      version_perk_policy: {VERSION_PERK_POLICY}")
    print("")

    print(f"[1/5] 准备 manifest: {manifest_path}")
    sqlite_path = ensure_sqlite_manifest(manifest_path, cache_dir)
    print(f"      SQLite: {sqlite_path}")

    print("[2/5] 读取 DestinyInventoryItemDefinition ...")
    items = load_inventory_items(sqlite_path)
    plug_sets = load_plug_sets(sqlite_path)
    index = ManifestIndex(items, plug_sets)
    print(f"      已读取 {len(items)} 个 InventoryItem")
    print(f"      已读取 {len(plug_sets)} 个 PlugSet")

    print(f"[3/5] 读取推荐表: {input_path}")
    headers, rows = read_table(input_path, sheet=SHEET_NAME)
    weapon_col, perk_cols, note_col, tier_col, rank_col = detect_columns(headers)
    print(f"      武器列: {weapon_col}")
    print(f"      Perk列: {', '.join(perk_cols)}")

    expanded = expand_rows(rows, weapon_col, perk_cols, note_col, tier_col, rank_col)
    print(f"      展开后 roll 组合: {len(expanded)}")
    write_extracted_report(extracted_path, expanded)
    print(f"      识别检查文件: {extracted_path}")
    write_weapon_candidate_report(weapon_candidates_path, index, expanded)
    print(f"      武器候选检查文件: {weapon_candidates_path}")
    write_perk_candidate_report(perk_candidates_path, index, expanded)
    print(f"      Perk候选检查文件: {perk_candidates_path}")

    print("[4/5] 匹配 hash 并生成 DIM wishlist ...")
    lines, unresolved, resolved_audit = build_wishlist(index, expanded)
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    csv_write(unresolved_path, unresolved, [
        "source_row", "type", "name", "reason", "original_reason", "weapon", "weapon_hash", "slot", "socket_index",
        "mapping_method", "mapping_name_hits", "generated", "version_status",
        "requested_option_count", "supported_option_count", "perks",
        "same_name_candidate_hashes", "same_name_candidate_types",
        "socket_candidate_hashes", "socket_candidate_names",
    ])
    csv_write(resolved_audit_path, resolved_audit, [
        "source_row", "weapon", "weapon_hash", "weapon_sqlite_id", "manifest_name",
        "version_status", "included_slots", "dropped_perks", "requested_option_count", "supported_option_count",
        "slot_1_barrel_requested_options", "slot_1_barrel_supported_options",
        "slot_1_barrel", "slot_1_barrel_hash", "slot_1_barrel_sqlite_id", "slot_1_barrel_type", "slot_1_barrel_socket_index",
        "slot_2_magazine_requested_options", "slot_2_magazine_supported_options",
        "slot_2_magazine", "slot_2_magazine_hash", "slot_2_magazine_sqlite_id", "slot_2_magazine_type", "slot_2_magazine_socket_index",
        "slot_3_trait_requested_options", "slot_3_trait_supported_options",
        "slot_3_trait", "slot_3_trait_hash", "slot_3_trait_sqlite_id", "slot_3_trait_type", "slot_3_trait_socket_index",
        "slot_4_trait_requested_options", "slot_4_trait_supported_options",
        "slot_4_trait", "slot_4_trait_hash", "slot_4_trait_sqlite_id", "slot_4_trait_type", "slot_4_trait_socket_index",
        "dim_line",
    ])

    resolved_count = sum(1 for line in lines if line.startswith("dimwishlist:"))
    print("[5/5] 完成")
    print(f"      resolved lines: {resolved_count}")
    print(f"      unresolved rows: {len(unresolved)}")
    print(f"      wishlist: {out_path}")
    print(f"      unresolved: {unresolved_path}")
    print(f"      extracted: {extracted_path}")
    print(f"      resolved audit: {resolved_audit_path}")
    print(f"      weapon candidates: {weapon_candidates_path}")
    print(f"      perk candidates: {perk_candidates_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
