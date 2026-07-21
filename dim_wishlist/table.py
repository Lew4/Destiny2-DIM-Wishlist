"""Recommendation-table parsing and Cartesian expansion."""

from __future__ import annotations

import csv
import itertools
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from .config import (
    EXCLUDE_PERK_HEADERS,
    NOTE_COL_ALIASES,
    RANK_COL_ALIASES,
    TIER_COL_ALIASES,
    WEAPON_COL_ALIASES,
)
from .utils import clean_text, norm_name, split_options

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - CSV users do not need openpyxl
    load_workbook = None


def make_unique_headers(raw_headers: Sequence[Any]) -> List[str]:
    """Preserve repeated headers such as the first two Perk columns."""
    counts: Dict[str, int] = {}
    headers = []
    for index, header in enumerate(raw_headers):
        base = clean_text(header) or f"col_{index + 1}"
        counts[base] = counts.get(base, 0) + 1
        headers.append(base if counts[base] == 1 else f"{base}__{counts[base]}")
    return headers


def looks_like_header(row: Sequence[Any]) -> bool:
    values = [clean_text(value) for value in row]
    aliases = {norm_name(alias) for alias in WEAPON_COL_ALIASES}
    return any(norm_name(value) in aliases for value in values) and any(
        "perk" in value.lower() for value in values
    )


def _build_from_rows(all_rows: List[List[Any]]) -> Tuple[List[str], List[Dict[str, Any]]]:
    header_index = next(
        (index for index, row in enumerate(all_rows[:80]) if looks_like_header(row)),
        None,
    )
    if header_index is None:
        preview = [[clean_text(value) for value in row] for row in all_rows[:8]]
        raise RuntimeError(
            "没有识别到表头。请确保某一行同时包含 `名字` 列和至少一个 `Perk` 列。"
            f"\n前几行内容: {preview}"
        )

    headers = make_unique_headers(all_rows[header_index])
    rows = []
    for raw in all_rows[header_index + 1:]:
        if not any(clean_text(value) for value in raw) or looks_like_header(raw):
            continue
        rows.append({
            header: raw[index] if index < len(raw) else None
            for index, header in enumerate(headers)
        })
    return headers, rows


def read_table(input_path: Path, sheet: Optional[str] = None) -> Tuple[List[str], List[Dict[str, Any]]]:
    suffix = input_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        if load_workbook is None:
            raise RuntimeError("读取 xlsx 需要 openpyxl：python3 -m pip install openpyxl")
        workbook = load_workbook(str(input_path), data_only=True)
        worksheet = workbook[sheet] if sheet else workbook.active
        return _build_from_rows([list(row) for row in worksheet.iter_rows(values_only=True)])
    if suffix == ".csv":
        with input_path.open("r", encoding="utf-8-sig", newline="") as stream:
            return _build_from_rows(list(csv.reader(stream)))
    raise RuntimeError(f"暂不支持输入格式: {suffix}")


def detect_columns(
    headers: Sequence[str],
) -> Tuple[str, List[str], Optional[str], Optional[str], Optional[str]]:
    weapon_aliases = {norm_name(alias) for alias in WEAPON_COL_ALIASES}
    weapon_col = next((header for header in headers if norm_name(header) in weapon_aliases), None)
    if not weapon_col:
        raise RuntimeError(f"没有找到武器名列。可用列名: {headers}")

    excluded = {norm_name(header) for header in EXCLUDE_PERK_HEADERS}
    perk_cols = []
    for header in headers:
        base = header.split("__", 1)[0]
        if norm_name(base) not in excluded and "perk" in base.lower():
            perk_cols.append(header)
    if not perk_cols:
        raise RuntimeError("没有找到 Perk 列。列名需要包含 `Perk`。")

    def optional_column(aliases: set[str]) -> Optional[str]:
        normalized = {norm_name(alias) for alias in aliases}
        return next((header for header in headers if norm_name(header) in normalized), None)

    return (
        weapon_col,
        perk_cols,
        optional_column(NOTE_COL_ALIASES),
        optional_column(TIER_COL_ALIASES),
        optional_column(RANK_COL_ALIASES),
    )


def make_note(
    row: Dict[str, Any],
    note_col: Optional[str],
    tier_col: Optional[str],
    rank_col: Optional[str],
) -> str:
    parts = []
    if tier_col and clean_text(row.get(tier_col)):
        parts.append(f"Tier {clean_text(row.get(tier_col))}")
    if rank_col and clean_text(row.get(rank_col)):
        parts.append(f"Rank {clean_text(row.get(rank_col))}")
    if note_col and clean_text(row.get(note_col)):
        parts.append(clean_text(row.get(note_col)).replace("\n", " "))
    return " | ".join(parts)


def canonical_perk_slot(column: str, index: int) -> str:
    base = column.split("__", 1)[0].strip().lower()
    if index == 0:
        return "slot_1_barrel"
    if index == 1:
        return "slot_2_magazine"
    if base in {"perk 1", "trait 1", "特性1", "特性 1"}:
        return "slot_3_trait"
    if base in {"perk 2", "trait 2", "特性2", "特性 2"}:
        return "slot_4_trait"
    return f"slot_{index + 1}"


def expand_rows(
    rows: List[Dict[str, Any]],
    weapon_col: str,
    perk_cols: List[str],
    note_col: Optional[str],
    tier_col: Optional[str],
    rank_col: Optional[str],
) -> List[Dict[str, Any]]:
    expanded = []
    for row_index, row in enumerate(rows, start=1):
        weapon = clean_text(row.get(weapon_col))
        if not weapon:
            continue
        option_lists: List[List[str]] = []
        slot_order: List[str] = []
        raw_columns: Dict[str, str] = {}
        parsed_columns: Dict[str, List[str]] = {}
        for column_index, column in enumerate(perk_cols):
            raw_value = clean_text(row.get(column))
            options = split_options(row.get(column))
            slot = canonical_perk_slot(column, column_index)
            if raw_value:
                raw_columns[slot] = raw_value
            if options:
                parsed_columns[slot] = options
                option_lists.append(options)
                slot_order.append(slot)
        if not option_lists:
            continue
        for combination in itertools.product(*option_lists):
            perk_slots = dict(zip(slot_order, combination))
            expanded.append({
                "source_row": row_index,
                "weapon": weapon,
                "perks": [perk_slots[slot] for slot in slot_order],
                "perk_slots": perk_slots,
                "slot_order": list(slot_order),
                "notes": make_note(row, note_col, tier_col, rank_col),
                "raw_perk_columns": raw_columns,
                "parsed_perk_columns": parsed_columns,
            })
    return expanded


def collapse_expanded_recommendations(expanded: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Restore one record per source recommendation before version fallback."""
    unique: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
    for record in expanded:
        parsed = record.get("parsed_perk_columns", {}) or {}
        signature = tuple(
            (slot, tuple(parsed.get(slot, []) or []))
            for slot in record.get("slot_order", []) or []
        )
        key = record.get("source_row"), record.get("weapon"), record.get("notes"), signature
        unique.setdefault(key, {
            "source_row": record.get("source_row"),
            "weapon": record.get("weapon", ""),
            "notes": record.get("notes", ""),
            "slot_order": list(record.get("slot_order", []) or []),
            "parsed_perk_columns": {
                slot: list(parsed.get(slot, []) or [])
                for slot in record.get("slot_order", []) or []
            },
        })
    return list(unique.values())
